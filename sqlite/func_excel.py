############################# HEADER ######################################
#
# beri obstojeco sqlite bazo in zapisi .xlsx datoteko
#
###########################################################################

# imports
from sqlite.opis import ExOpis

from re import sub, split
from openpyxl import Workbook, styles

class ExDatoteka:

    def __init__(self, projekt='popis'):
        ExDatoteka.projekt = projekt
        ExDatoteka.zvezek = Workbook()

    @classmethod
    def ex_shrani(self):
        # self.ex_zbrisi_sheet()
        ExDatoteka.zvezek.save(ExDatoteka.projekt+'.xlsx')
    
    @classmethod
    def ex_zbrisi_sheet(self):
    
        std = ExDatoteka.zvezek.get_sheet_by_name('Sheet')
        ExDatoteka.zvezek.remove_sheet(std)


class ExStran(ExDatoteka):

    def __init__(self, datoteka):
        zvezek = datoteka.zvezek
        self.ex_stran = zvezek['Sheet']
        self.zap_popis = 1
        self.r_kurzor = 3

    def temp_naslovna_vr(self):
        """
        Vpis vrednosti v naslovno vrstico
        """
        naslovna = ['Zap. št.', 'Prodajni program', 'Količina',
                    'Cena/kos', 'Prodajna cena']
        for stolpec, vrednost in enumerate(naslovna, 1):
            _ = self.ex_stran.cell(column=stolpec, row=1, value=vrednost)

    def temp_dimenzioniraj(self):
        """
        Oblikovanje naslovne vrstice
        """
        self.ex_stran.column_dimensions['A'].width = 5
        self.ex_stran.column_dimensions['B'].width = 60
        self.ex_stran.column_dimensions['C'].width = 10
        self.ex_stran.column_dimensions['D'].width = 15
        self.ex_stran.column_dimensions['E'].width = 15
        double = styles.Side(border_style="double", color="111111")
        for c in self.ex_stran[1]:
            c.fill = styles.PatternFill("solid", fgColor='ffff99')
            c.alignment = styles.Alignment(wrap_text=True)
            c.border = styles.Border(bottom=double)
        for v in self.ex_stran.iter_rows():
            v[1].alignment = styles.Alignment(wrap_text=True)
            if v[0].value:
                for c in v:
                    c.font = styles.Font(bold=True)
            elif v[2].value:
                for c in v[3:]:
                    c.number_format = '0.00'

    def zapri(self):
        self.temp_naslovna_vr()
        self.temp_dimenzioniraj()

    def zapisi_postavko(self, productID, kolicina=1, cena=1):
        opis = ExOpis(productID)
        # vpis zaporedne stevilke
        self.ex_stran.cell(column=1, row=self.r_kurzor, 
            value=f'{self.zap_popis}.')
        # ustvarimo celice in vstavimo podatke postavke
        for i in range(len(opis.podatki)):
            self.ex_stran.cell(column=2, row=self.r_kurzor, 
                value=opis.podatki[i])
            self.r_kurzor += 1
        # vpis kolicine
        self.ex_stran.cell(column=3, row=self.r_kurzor - 1, 
            value = int(kolicina))
        # vpis cene/kos
        self.ex_stran.cell(column=4, row=self.r_kurzor - 1, 
            value = int(cena))
        # cena
        self.ex_stran.cell(column=5, row=self.r_kurzor - 1, 
            value = int(kolicina)*int(cena))
        # ustvarjenim celicam dodamo oblikovanje
        # self.temp_krepko()
        # premaknemo kurzor za vstavljanje nove postavke
        self.r_kurzor += 2
        self.zap_popis += 1
        

# def main2():
#     zvezek = ExDatoteka()
#     stran = ExStran(zvezek)

#     stran.zapisi_postavko('i-BX M 004M')
#     stran.zapisi_postavko('i-BX T 035T')
#     stran.zapri()

#     zvezek.ex_shrani()

# main2()
    