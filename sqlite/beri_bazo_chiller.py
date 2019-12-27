############################# HEADER ######################################
#
# beri obstojeco sqlite bazo in zapisi .xlsx datoteko
#
###########################################################################

# imports
from re import sub, split
import sqlite3
from openpyxl import Workbook, styles


# region VpisExcel
# class Baza():

#     def __init__(self, segment='chiller', skupine=None, izvedbe = None, velikosti = None, velikosti_cevni = None):
#         self.segment = segment
#         self.skupine = skupine
#         self.izvedbe = izvedbe
#         self.velikosti = velikosti
#         self.velikosti_cevni = velikosti_cevni

segment='chiller'
skupine=None
izvedbe = None
velikosti = None
velikosti_cevni = None

def poisci_skupine():
    global skupine
    b = sqlite3.connect("SQL_"+segment+".db")
    baza = b.cursor()
    baza.execute('''SELECT DISTINCT "{}" 
                    FROM "GENERAL" 
                    ORDER BY NumID'''.format('Group'))
    skupine = [i[0] for i in baza.fetchall()]
    b.close()
    return skupine

def poisci_izvedbe(skupina):
    global izvedbe
    b = sqlite3.connect("SQL_"+segment+".db")
    baza = b.cursor()
    baza.execute('''SELECT DISTINCT "{}" 
                    FROM "GENERAL"
                    WHERE "Group"="{}"
                    ORDER BY NumID'''.format('Version', skupina))
    izvedbe = [i[0] for i in baza.fetchall()]
    b.close()
    return izvedbe

def poisci_velikosti(skupina, izvedba):
    global velikosti
    b = sqlite3.connect("SQL_"+segment+".db")
    baza = b.cursor()
    baza.execute('''SELECT DISTINCT "{}" 
                    FROM "GENERAL"
                    WHERE "Group"="{}" AND Version="{}"
                    ORDER BY NumID'''.format('Size', skupina, izvedba))
    velikosti = [i[0] for i in baza.fetchall() if i[-1] != 'T']
    b.close()
    print(velikosti)
    return velikosti

def poisci_velikosti_cevni(skupina, izvedba):
    global velikosti_cevni
    b = sqlite3.connect("SQL_"+segment+".db")
    baza = b.cursor()
    baza.execute('''SELECT DISTINCT "{}" 
                    FROM "GENERAL"
                    WHERE "Group"="{}" AND Version={}
                    ORDER BY NumID'''.format('Size', skupina, izvedba))
    velikosti_cevni = [i[0] for i in baza.fetchall() if i[-1] == 'T']
    b.close()
    return velikosti_cevni

class ExDatoteka:
    st_opis = 0
    st_list = 0
    st_zvezek = 0
    zvezek = Workbook()
    ex_skupina = 'Neznano'
    ex_izvedbe = set()

    def __init__(self, ex_skupina):
        ExDatoteka.ex_skupina = ex_skupina
        ExDatoteka.st_opis = 0
        ExDatoteka.st_zvezek += 1
        ExDatoteka.zvezek = Workbook()

    @classmethod
    def ex_shrani(self):
        ExDatoteka.zvezek.save(self.ex_skupina + '.xlsx')
    
    @classmethod
    def ex_zbrisi_sheet(self):
        std = ExDatoteka.zvezek.get_sheet_by_name('Sheet')
        ExDatoteka.zvezek.remove_sheet(std)

class ExStran(ExDatoteka):

    ex_opis = '***Splošni opis enote***'
    # postavke = [('Hladilna moč', 'kW'), ('EER (EN14511 metoda)', ''), 
    #     ('ESEER (EN14511 metoda)', ''), ('SEER (Reg. EU 2016/2281)', ''), ('Električna moč', 'kW'), ('El. priključek', ''), 
    #     ('Zvočni tlak (SPL)', 'dB(A)'), ('Zvočna moč (PWL)', 'dB(A)'),
    #     ('Število hladilnih krogov', ''), ('Število kompresorjev', ''), ('Dolžina', 'mm'), ('Širina', 'mm'),('Višina', 'mm'), ('Teža', 'kg')]
    
    postavke = [('Hladilna moč', 'kW'), ('EER (EN14511 metoda)', ''), 
        ('ESEER (EN14511 metoda)', ''), ('SEER (Reg. EU 2016/2281)', ''),
         ('El. priključek', ''), 
        ('Zvočni tlak (SPL)', 'dB(A)'), ('Zvočna moč (PWL)', 'dB(A)'),
        ('Število hladilnih krogov', ''), ('Število kompresorjev', ''), ('Dolžina', 'mm'), ('Širina', 'mm'),('Višina', 'mm'), ('Teža', 'kg')]

    def __init__(self, ex_izvedba):
        self.ex_izvedba = ex_izvedba
        
        if  not ex_izvedba in ExDatoteka.zvezek.sheetnames:
            temp_ime_lista = '_'.join(
                [ExDatoteka.ex_skupina, ex_izvedba])
            ExDatoteka.zvezek.create_sheet(temp_ime_lista)
            self.ex_stran = ExDatoteka.zvezek[temp_ime_lista]

        ExDatoteka.ex_izvedbe.add(ex_izvedba)
        ExDatoteka.st_list += 1
        ExDatoteka.st_opis = 0
    
    def temp_naslovna_vr(self):
        naslovna = ['Zap. št.', 'Prodajni program', 'Količina',
                    'Cena/kos', 'Prodajna cena']
        for stolpec, vrednost in enumerate(naslovna, 1):
            _ = self.ex_stran.cell(column=stolpec, row=1, value=vrednost)

    def temp_dimenzioniraj(self):
        self.ex_stran.column_dimensions['A'].width = 5
        self.ex_stran.column_dimensions['B'].width = 60
        self.ex_stran.column_dimensions['C'].width = 10
        self.ex_stran.column_dimensions['D'].width = 15
        self.ex_stran.column_dimensions['E'].width = 15
        double = styles.Side(border_style="double", color="111111")
        for c in self.ex_stran[1]:
            c.fill = styles.PatternFill("solid", fgColor='ffff99')
            c.alignment = styles.Alignment(wrap_text=True)
            c.border = styles.Border(bottom=double)
        for v in self.ex_stran.iter_rows():
            v[1].alignment = styles.Alignment(wrap_text=True)
            if v[0].value:
                for c in v:
                    c.font = styles.Font(bold=True)
            elif v[2].value:
                for c in v[3:]:
                    c.number_format = '0.00'

    def ex_zapisi_podatke(self, objekt):
        i = objekt.t_dol * (ExDatoteka.st_opis-1) + 3
        self.ex_stran.cell(column=1, row=i, 
            value=str(ExDatoteka.st_opis) + '.')
        self.ex_stran.cell(column=2, row=i, 
            value=' '.join(['Hladilni agregat Climaveneta', ExDatoteka.ex_skupina + '/', self.ex_izvedba, objekt.velikost]))
        self.ex_stran.cell(column=2, row=i+1, value=objekt.ex_opis)
        self.ex_stran.cell(column=2, row=i+2, 
            value=' '.join(
                ['PROIZVAJALEC:', 
                'Mitsubishi Electric Hydronics & IT Cooling Systems S.p.A, Italija']))
        self.ex_stran.cell(column=2, row=i+3, 
            value='UVOZNIK: REAM d.o.o., Trzin')
        self.ex_stran.cell(column=2, row=i+5, value='TEHNIČNI OPIS:')
        for j in range(len(objekt.tehnicni_podatki)):
            if objekt.tehnicni_podatki[j]:
                _ = '{}:    {} {}'.format(ExStran.postavke[j][0], 
                    objekt.tehnicni_podatki[j], ExStran.postavke[j][1])
                self.ex_stran.cell(column=2, row=i+j+6, value=_)
        vr = i+len(objekt.tehnicni_podatki)+5
        self.ex_stran.cell(column=3, row=vr, value=1)
        self.ex_stran.cell(column=4, row=vr, value=0.00)
        self.ex_stran.cell(column=5, row=vr, 
            value=r'=$C{}*$D{}'.format(vr, vr))
        return None

class ExOpis(ExStran):

    def __init__(self, ob_stran, velikost):
        ExDatoteka.st_opis += 1
        self.velikost = velikost
        self.ex_productID = ' '.join(
            filter(None, 
            [ExDatoteka.ex_skupina, ob_stran.ex_izvedba, self.velikost]))
        self.tehnicni_podatki = []
        self.t_dol = 0

    def ex_sestavi_naziv(self):
        return ' '.join(['Hladilni agragat Climaveneta', self.ex_productID])
    
    def ex_dimenzije(self):
        b = sqlite3.connect("SQL_"+segment+".db")
        baza = b.cursor()
        baza.execute('''SELECT "A", "B", "H" 
                        FROM GENERAL
                        WHERE productID=?''', (self.ex_productID,))
        _ = [i for i in baza.fetchall()[0]]
        dimenzije = '{} x {} x {}'.format(*_)
        b.close()
        return dimenzije

    def ex_teh_opis(self):
        '''
        TEHNIČNI PODATKI:
        Hladilna moč:
        EER (EN14511 metoda):
        ESEER (EN14511 metoda):
        SEER (Reg. EU 2016/2281):
        El.moč:
        El. priključek:
        Zvočni tlak SPL:
        Zvočna moč PWL:
        Število hladilnih krogov:
        Število kompresorjev:
        Dolžina:
        Širina:
        Višina:
        Teža:
        '''

        b = sqlite3.connect("SQL_"+segment+".db")
        baza = b.cursor()
        baza.execute('''SELECT "Cooling capacity" 
                        FROM COOLING_EUROVENT
                        WHERE productID=?''', (self.ex_productID,))
        self.tehnicni_podatki.append(baza.fetchall()[0][0])
        baza.execute('''SELECT "EER", "ESEER" 
                        FROM COOLING_EUROVENT
                        WHERE productID=?''', (self.ex_productID,))
        for i in baza.fetchall()[0]:
            try:
                ba, de = i.split(',')
                self.tehnicni_podatki.append(ba+','+de[:2])
            except:
                if i:
                    self.tehnicni_podatki.append(i)
                else:
                    self.tehnicni_podatki.append('-')
        baza.execute('''SELECT "SEER"
                        FROM SEASONAL_EFF_COOLING
                        WHERE productID=?''', (self.ex_productID,))
        i = baza.fetchall()[0][0].split(',')
        try:
            ba, de = i
            self.tehnicni_podatki.append(ba+','+de[:2])
        except:
            self.tehnicni_podatki.append('-')
        baza.execute('''SELECT "Total power input" 
                        FROM COOLING_GROSS
                        WHERE productID=?''', (self.ex_productID,))
        # self.tehnicni_podatki.append(baza.fetchall()[0][0])
        baza.execute('''SELECT "Power supply"
                        FROM GENERAL
                        WHERE productID=?''', (self.ex_productID,))
        if '400' in baza.fetchall()[0][0]:
            self.tehnicni_podatki.append('400V/ 3F/ 50Hz')
        else:
            self.tehnicni_podatki.append('230V/ 1F/ 50Hz')
        baza.execute('''SELECT "Sound Pressure", 
                        "Sound power level in cooling", "No. Circuits", "Compressors nr.", "A", "B", "H", "Operating weight"
                        FROM GENERAL
                        WHERE productID=?''', (self.ex_productID,))
        self.tehnicni_podatki.extend([i for i in baza.fetchall()[0]])
        b.close()
        self.t_dol = len(list(filter(None,self.tehnicni_podatki))) + 8
        print(self.tehnicni_podatki)
        return self.tehnicni_podatki


def main2():
    poisci_skupine()
    for sk in skupine:
        dat = ExDatoteka(sk)
        # datoteka se nanasa na skupini agregatov
        for iz in poisci_izvedbe(sk):
            stran = ExStran(iz)
            # dat.zamenjaj_list()
            # dodaj stil strani
            for vel in poisci_velikosti(sk, iz):
                opis = ExOpis(stran, vel)
                opis.ex_teh_opis()
                stran.ex_zapisi_podatke(opis)
                # print(opis.tehnicni_podatki)
            stran.temp_naslovna_vr()
            stran.temp_dimenzioniraj()
        dat.ex_zbrisi_sheet()
        dat.ex_shrani()
main2()
    
# endregion